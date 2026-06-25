import os
import pandas as pd
import openpyxl

# ==========================================
# 1. CONFIGURAÇÃO DE CAMINHOS (Ajuste aqui!)
# ==========================================
CAMINHO_REDE = "dados"

ARQUIVO_ENTRADA = "dados_importacao_simulados.xlsx"
ARQUIVO_TEMPLATE = "Template_Dashboard.xlsx"
ARQUIVO_SAIDA = "Dashboard_Teste_Diretoria.xlsx" # O que os gestores vão abrir

# Monta os caminhos completos
path_bruto = os.path.join(CAMINHO_REDE, ARQUIVO_ENTRADA)
path_template = os.path.join(CAMINHO_REDE, ARQUIVO_TEMPLATE)
path_final = os.path.join(CAMINHO_REDE, ARQUIVO_SAIDA)

print("🚀 Iniciando processamento do Dashboard...")

# ==========================================
# 2. LEITURA E EXTRAÇÃO DA BASE MÃE
# ==========================================
print("📥 Lendo base bruta do Excel...")
df = pd.read_excel(path_bruto, sheet_name=0) # Primeira aba

# Tratar datas e criar colunas de Ano e Mês
df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, errors='coerce')
df['ANO'] = df['Data'].dt.year
df['MÊS'] = df['Data'].dt.month

# ==========================================
# 3. FILTROS INTELIGENTES (CÁLCULOS DO PYTHON)
# ==========================================
print("📊 Calculando o Top 5 de Concorrentes...")
produtos_alvo = ["ALVEOLAR", "COMPACTA"]
df_2026_filtrado = df[(df['ANO'] == 2026) & (df['ID_ITEM'].isin(produtos_alvo))]

todas_empresas_top = []

for prod in produtos_alvo:
    df_prod = df_2026_filtrado[df_2026_filtrado['ID_ITEM'] == prod]
    top_5 = df_prod.groupby('Notificado_Final')['VALOR FOB ESTIMADO TOTAL'].sum().nlargest(5).index.tolist()
    todas_empresas_top.extend(top_5)

empresas_finais_unicas = list(set(todas_empresas_top))

# Filtrar base histórica (2025 e 2026) e agrupar
df_historico = df[
    (df['Notificado_Final'].isin(empresas_finais_unicas)) & 
    (df['ANO'].isin([2025, 2026])) &
    (df['ID_ITEM'].isin(produtos_alvo))
]

tabela_para_excel = (
    df_historico.groupby(['ID_ITEM', 'Notificado_Final', 'ANO', 'MÊS'])['VALOR FOB ESTIMADO TOTAL']
    .sum()
    .reset_index()
)

# ==========================================
# 4. INJETANDO OS DADOS NO TEMPLATE DO EXCEL
# ==========================================
print("📝 Gravando dados no arquivo Excel dos gestores...")

# Abre o arquivo modelo (Template)
wb = openpyxl.load_workbook(path_template)

# Seleciona a aba que vai receber os dados brutos do Python
if "Dados_Python" in wb.sheetnames:
    ws = wb["Dados_Python"]
    # Limpa dados antigos se houver
    ws.delete_rows(1, ws.max_row+1)
else:
    ws = wb.create_sheet("Dados_Python")

# Escreve o cabeçalho das colunas
ws.append(list(tabela_para_excel.columns))

# Injeta todas as linhas calculadas do Pandas para dentro das células do Excel
for index, row in tabela_para_excel.iterrows():
    ws.append(list(row))

# Salva como um novo arquivo oficial (para não estragar o seu template)
wb.save(path_final)

print(f"✅ Sucesso! Dashboard atualizado salvo em: {path_final}")
